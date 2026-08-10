#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Importer — reads patient names from .xlsx files.

Scans all tabs, detects the patient column from headers (PACIENTE(S),
NOME(S), NAME(S), ...) and extracts unique normalized names.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date

from andaime.error_handler import ErrorContext, ErrorHandler, ErrorLevel
from andaime.text import to_upper_normalized

_HEADER_SCAN_ROWS = 16
_CONTENT_SCAN_ROWS = 50
_MAX_EXTRACT_ROWS = 200_000
_TRAILING_BLANK_STOP = 8
_HIGH_PRIORITY = {"PACIENTE", "PACIENTES", "PATIENT", "PATIENTS"}
_LOW_PRIORITY = {"NOME", "NOMES", "NAME", "NAMES"}
_MED_PRIORITY = {"MEDICAMENTO", "MEDICAMENTOS", "MED", "MEDICACAO", "ITEM", "ITENS"}
_HEADER_NOISE = frozenset(_HIGH_PRIORITY | _LOW_PRIORITY)
_TOKEN_RE = re.compile(r"[A-Z0-9]+")

_COL_SPEC_RE = re.compile(r"^(\d+)\s*-\s*([A-Za-z]+)$")
_MALOTE_COL_SPEC_RE = re.compile(r"^(\d+)-([A-Za-z]+)-([A-Za-z]+)$")

# Tipo names found in tab titles (normalized without accents).
_TIPO_TITLES = {
    "ABERTURA": "entrada",
    "ENTRADA": "entrada",
    "ENTRADAS": "entrada",
    "RENOVACAO": "renovacao",
    "RENOVACOES": "renovacao",
    "RETIRADA": "retirada",
    "RETIRADAS": "retirada",
    "RESOLVER NA HORA": "urgente",
    "URGENTE": "urgente",
    "MEDICAMENTO EM CASA": "medcasa",
    "REMEDIO EM CASA": "medcasa",
    "MEDCASA": "medcasa",
}

# Tab names that map directly to a tipo when a title match is absent.
_TAB_TIPO = {
    "ABERTURA": "entrada",
    "ENTRADA": "entrada",
    "RENOVACAO": "renovacao",
    "RETIRADA": "retirada",
    "RESOLVER NA HORA": "urgente",
}

_MED_SPLIT_RE = re.compile(r"\s*;\s*|\s+/\s+|\s*-\s*")
_DATE_TOKENS_RE = re.compile(r"\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?")


def _header_tokens(text: str) -> set[str]:
    return set(_TOKEN_RE.findall(to_upper_normalized(text)))


def _nonempty_cell_count(row: tuple) -> int:
    return sum(1 for v in row if v is not None and str(v).strip())


def parse_malote_date(text: str) -> date | None:
    """Parse 'DD.MM.YYYY' or 'DD.MM' (separators / . -).

    A year-less date (DD.MM) uses the current year.
    """
    for sep in ("/", "-", "."):
        if sep in text:
            parts = [p for p in text.split(sep) if p.strip()]
            break
    else:
        return None
    if len(parts) not in (2, 3):
        return None
    try:
        day = int(parts[0])
        month = int(parts[1])
        year = None if len(parts) == 2 else int(parts[2])
    except ValueError:
        return None
    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None
    if year is None:
        year = date.today().year
    elif year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def split_medications(text: str) -> list[str]:
    """Split a medications cell into individual med names.

    Separators: ';' (always), '/' with spaces both sides (' / '), and '-'
    when a space touches it. A glued 'MG/ML' or 'ML-KIT' stays intact.
    """
    parts = _MED_SPLIT_RE.split(text)
    return [p.strip() for p in parts if p.strip()]


def parse_col_spec(text: str) -> dict[int, str]:
    """Parse '1-A; 2-B' into {1: 'A', 2: 'B'}.

    Each entry is ``<tab_index>-<column_letter>`` separated by ``;``.
    Malformed entries are ignored. Returns a dict keyed by 1-based tab index.
    """
    result: dict[int, str] = {}
    for part in text.split(";"):
        part = part.strip()
        if not part:
            continue
        m = _COL_SPEC_RE.match(part)
        if not m:
            continue
        idx = int(m.group(1))
        if idx >= 1:
            result[idx] = m.group(2).upper()
    return result


def format_col_spec(col_map: dict[int, str]) -> str:
    """Inverse of parse_col_spec: {1: 'A', 2: 'B'} -> '1-A; 2-B'."""
    return "; ".join(f"{idx}-{letter}" for idx, letter in sorted(col_map.items()))


def parse_malote_spec(text: str) -> dict[int, tuple[str, str]]:
    """Parse '1-A-B; 2-A-C' into {1: ('A','B'), 2: ('A','C')}.

    Each entry is ``<tab>-<patient_col>-<med_col>``. Malformed entries ignored.
    """
    result: dict[int, tuple[str, str]] = {}
    for part in text.split(";"):
        part = part.strip()
        if not part:
            continue
        m = _MALOTE_COL_SPEC_RE.match(part)
        if not m:
            continue
        idx = int(m.group(1))
        if idx >= 1:
            result[idx] = (m.group(2).upper(), m.group(3).upper())
    return result


def _ensure_openpyxl():
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        ErrorHandler.log(
            "openpyxl não instalado",
            level=ErrorLevel.ERROR,
            context=ErrorContext.EXPORT,
        )
        return None


@dataclass
class SheetAnalysis:
    name: str
    content_columns: list[str]
    header_row: int | None
    patient_col: str | None
    med_col: str | None = None
    tipo: str | None = None
    malote_date: date | None = None

    @property
    def is_malote_tab(self) -> bool:
        return bool(self.tipo and self.patient_col)


@dataclass
class ImportAnalysis:
    sheets: list[SheetAnalysis]

    @property
    def all_content_columns(self) -> list[str]:
        seen: list[str] = []
        for s in self.sheets:
            for c in s.content_columns:
                if c not in seen:
                    seen.append(c)
        return seen

    @property
    def detected_patient_col(self) -> str | None:
        for s in self.sheets:
            if s.patient_col:
                return s.patient_col
        return None

    @property
    def default_col_spec(self) -> str:
        """Pre-fill string from per-sheet detection, e.g. '1-A; 3-B'."""
        parts = [f"{i}-{s.patient_col}" for i, s in enumerate(self.sheets, start=1) if s.patient_col]
        return "; ".join(parts)

    @property
    def sheet_names(self) -> list[str]:
        return [s.name for s in self.sheets]

    @property
    def has_malote(self) -> bool:
        """True when at least one tab looks like a malote tab."""
        return any(s.is_malote_tab for s in self.sheets)

    @property
    def default_malote_spec(self) -> str:
        """Pre-fill 'tab-PAC-MED' spec from detection, e.g. '1-A-B; 2-A-B'."""
        parts = [f"{i}-{s.patient_col}-{s.med_col}" for i, s in enumerate(self.sheets, start=1) if s.is_malote_tab and s.med_col]
        return "; ".join(parts)


@dataclass
class MaloteRow:
    """One patient's meds for a single process row/group in a tab."""

    patient: str
    meds: list[str]


@dataclass
class MaloteTab:
    """Extracted data for a single malote tab (tipo)."""

    tab_index: int
    sheet_name: str
    tipo: str
    date: date | None
    rows: list[MaloteRow] = field(default_factory=list)


@dataclass
class MaloteExtraction:
    """Result of malote extraction across tabs."""

    tabs: list[MaloteTab] = field(default_factory=list)


class ExcelImporter:
    """Open a workbook once, analyze it, then extract names from a chosen column."""

    def __init__(self, path: str) -> None:
        openpyxl = _ensure_openpyxl()
        if openpyxl is None:
            raise ImportError("openpyxl não instalado")
        self.path = path
        self._wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self._analysis: ImportAnalysis | None = None

    def close(self) -> None:
        self._wb.close()

    def analyze(self) -> ImportAnalysis:
        sheets = [self._analyze_sheet(ws) for ws in self._wb.worksheets]
        self._analysis = ImportAnalysis(sheets=sheets)
        return self._analysis

    def _analyze_sheet(self, ws) -> SheetAnalysis:
        from openpyxl.utils import get_column_letter

        max_row = ws.max_row or 0
        if max_row == 0:
            return SheetAnalysis(
                name=ws.title, content_columns=[], header_row=None, patient_col=None
            )

        rows = list(
            ws.iter_rows(
                min_row=1,
                max_row=min(max_row, _CONTENT_SCAN_ROWS),
                values_only=True,
            )
        )
        if not rows:
            return SheetAnalysis(
                name=ws.title, content_columns=[], header_row=None, patient_col=None
            )

        n_cols = max(len(r) for r in rows)

        content_columns: list[str] = []
        for col_idx in range(1, n_cols + 1):
            for r in rows:
                if col_idx <= len(r):
                    v = r[col_idx - 1]
                    if v is not None and str(v).strip():
                        content_columns.append(get_column_letter(col_idx))
                        break

        # ---- malote metadata from the first rows (title area) ----
        tipo: str | None = _TAB_TIPO.get(to_upper_normalized(ws.title))
        malote_date: date | None = None
        title_text = ""
        for r in rows[:_HEADER_SCAN_ROWS]:
            for v in r:
                if not isinstance(v, str):
                    continue
                text = v.strip()
                if text and ("-" in text or "/" in text or "." in text):
                    title_text = f"{title_text}\n{text}".strip()
        for line in title_text.split("\n")[:5]:
            upper = to_upper_normalized(line)
            if not upper:
                continue
            if tipo is None:
                for token, mapped in _TIPO_TITLES.items():
                    if upper.startswith(token) or f" {token} " in f" {upper} ":
                        tipo = mapped
                        break
            if malote_date is None:
                malote_date = self._find_date_in_title(upper)
        if tipo is None:
            for token, mapped in _TIPO_TITLES.items():
                if token in to_upper_normalized(ws.title):
                    tipo = mapped
                    break

        # ---- header row: patient + med columns (independent tracks) ----
        header_row: int | None = None
        patient_col: str | None = None
        med_col: str | None = None
        best_score = 0
        for row_idx, r in enumerate(rows[:_HEADER_SCAN_ROWS], start=1):
            if _nonempty_cell_count(r) < 2:
                continue
            row_has_match = False
            for col_idx in range(1, len(r) + 1):
                tokens = _header_tokens(str(r[col_idx - 1] or ""))
                if tokens & _HIGH_PRIORITY:
                    score = 2
                    kind = "patient"
                elif tokens & _LOW_PRIORITY:
                    score = 1
                    kind = "patient"
                elif tokens & _MED_PRIORITY:
                    score = 1
                    kind = "med"
                else:
                    continue
                row_has_match = True
                if kind == "med":
                    if med_col is None:
                        med_col = get_column_letter(col_idx)
                else:
                    if score > best_score:
                        best_score = score
                        patient_col = get_column_letter(col_idx)
                        header_row = row_idx
            if row_has_match:
                break

        if patient_col is None and 2 <= len(content_columns) <= 3:
            patient_col = content_columns[0]

        return SheetAnalysis(
            name=ws.title,
            content_columns=content_columns,
            header_row=header_row,
            patient_col=patient_col,
            med_col=med_col,
            tipo=tipo,
            malote_date=malote_date,
        )

    def _find_date_in_title(self, upper_text: str) -> date | None:
        for token in _DATE_TOKENS_RE.findall(upper_text):
            parsed = parse_malote_date(token)
            if parsed is not None:
                return parsed
        return None

    def _iter_data_rows(self, ws, start_row: int, ref_cols: tuple[int, ...]):
        """Iterate data rows from ``start_row``, stopping at a trailing run of
        blank rows or a hard cap. Guards against sheets whose formatting extends
        max_row to 1M+ rows with no content."""
        blank = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True)):
            if row_idx >= _MAX_EXTRACT_ROWS:
                break
            values = [row[c] if c < len(row) else None for c in ref_cols]
            if all(v is None or not str(v).strip() for v in values):
                blank += 1
                if blank > _TRAILING_BLANK_STOP:
                    break
                continue
            blank = 0
            yield row

    def extract_names(self, col_map: dict[int, str]) -> list[str]:
        """Extract unique normalized patient names.

        col_map maps a 1-based tab index to a column letter. Tabs absent from
        the map are skipped.
        """
        from openpyxl.utils import column_index_from_string

        if self._analysis is None:
            self.analyze()

        assert self._analysis is not None
        analysis_by_idx = {
            i: s for i, s in enumerate(self._analysis.sheets, start=1)
        }

        worksheets = self._wb.worksheets
        names: list[str] = []
        seen: set[str] = set()
        for tab_idx, col_letter in col_map.items():
            if tab_idx < 1 or tab_idx > len(worksheets):
                continue
            ws = worksheets[tab_idx - 1]
            col_idx = column_index_from_string(col_letter)
            sa = analysis_by_idx.get(tab_idx)
            start_row = (sa.header_row + 1) if sa and sa.header_row else 1
            for row in self._iter_data_rows(ws, start_row, (col_idx - 1,)):
                if col_idx - 1 >= len(row):
                    continue
                v = row[col_idx - 1]
                if v is None:
                    continue
                text = str(v).strip()
                if not text:
                    continue
                normalized = to_upper_normalized(text)
                if not normalized or normalized in _HEADER_NOISE:
                    continue
                if normalized not in seen:
                    seen.add(normalized)
                    names.append(normalized)
        return names

    def extract_malote(
        self, spec: dict[int, tuple[str, str]]
    ) -> MaloteExtraction:
        """Extract malote data from the given tabs.

        spec maps a 1-based tab index to (patient_col, med_col) letters.
        Rows are grouped by patient; each data row becomes one process group.
        """
        from openpyxl.utils import column_index_from_string

        if self._analysis is None:
            self.analyze()

        assert self._analysis is not None
        analysis_by_idx = {
            i: s for i, s in enumerate(self._analysis.sheets, start=1)
        }

        worksheets = self._wb.worksheets
        tabs: list[MaloteTab] = []
        for tab_idx, (pac_letter, med_letter) in spec.items():
            if tab_idx < 1 or tab_idx > len(worksheets):
                continue
            ws = worksheets[tab_idx - 1]
            sa = analysis_by_idx.get(tab_idx)
            pac_idx = column_index_from_string(pac_letter) - 1
            med_idx = column_index_from_string(med_letter) - 1
            start_row = (sa.header_row + 1) if sa and sa.header_row else 1

            rows: list[MaloteRow] = []
            blank = 0
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=start_row, values_only=True)
            ):
                if row_idx >= _MAX_EXTRACT_ROWS:
                    break
                pac = row[pac_idx] if pac_idx < len(row) else None
                med = row[med_idx] if med_idx < len(row) else None
                if (pac is None or not str(pac).strip()) and (
                    med is None or not str(med).strip()
                ):
                    blank += 1
                    if blank > _TRAILING_BLANK_STOP:
                        break
                    continue
                blank = 0
                patient = str(pac).strip() if pac is not None else ""
                if not patient or patient in _HEADER_NOISE:
                    continue
                meds = split_medications(str(med).strip()) if med else []
                rows.append(MaloteRow(patient=to_upper_normalized(patient), meds=meds))

            tabs.append(
                MaloteTab(
                    tab_index=tab_idx,
                    sheet_name=sa.name if sa else ws.title,
                    tipo=sa.tipo if sa and sa.tipo else "",
                    date=sa.malote_date if sa else None,
                    rows=rows,
                )
            )
        return MaloteExtraction(tabs=tabs)
