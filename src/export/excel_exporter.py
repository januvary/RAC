#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Exporter
Generates .xlsx spreadsheet from malote registros
"""

from __future__ import annotations

import re
from typing import TYPE_CHECKING, Optional
from datetime import datetime
from pathlib import Path

from andaime.error_handler import ErrorHandler, ErrorLevel

from src.constants import TIPO_LABELS, TIPO_TITLES

if TYPE_CHECKING:
    from src.database.rac_database import RACDatabase


class SavePathError(Exception):
    pass


def _format_item(name: str) -> str:
    paren = re.search(r"\(([^)]+)\)\s*$", name)
    if not paren:
        result = name
    else:
        brand = paren.group(1).strip().upper()
        digit = re.search(r"\d", name)
        if not digit:
            result = brand
        else:
            dosage = name[digit.start() : paren.start()].strip()
            result = f"{brand} {dosage}"
    return result.replace(" ", "\u00a0")


def _ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        ErrorHandler.log(
            "openpyxl não instalado",
            level=ErrorLevel.ERROR,
            context="Exportação",
        )
        return None


def _make_excel_styles():
    from openpyxl.styles import Alignment, Border, Font, Side

    return {
        "main_font": Font(name="Arial", size=11),
        "title1_font": Font(name="Arial", size=20),
        "title2_font": Font(name="Arial", size=16),
        "center": Alignment(horizontal="center", vertical="center"),
        "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "thin_border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def _apply_page_setup(ws):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.pageOrder = "downThenOver"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def _style_title_row(ws, row_num, styles, font_key="title1_font", height=30):
    for cell in ws[row_num]:
        cell.font = styles[font_key]
        cell.alignment = styles["center"]
        cell.border = styles["thin_border"]
    ws.row_dimensions[row_num].height = height


def _style_data_rows(ws, start_row, styles):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = styles["main_font"]
            cell.alignment = styles["left_wrap"]
            if cell.value is not None:
                cell.border = styles["thin_border"]


class ExcelExporter:
    def __init__(self, db: RACDatabase) -> None:
        self._db = db

    def _save_workbook(self, wb, base_filename, date_label="", log_label="Planilha exportada"):
        try:
            timestamp = datetime.now().strftime("%H%M%S")
            if date_label:
                filename = f"{base_filename}_{date_label}_{timestamp}.xlsx"
            else:
                filename = f"{base_filename}_{timestamp}.xlsx"

            from andaime.config import ConfigManager

            config = ConfigManager()
            save_path = config.get("save_path", Path.home() / "Downloads")

            if isinstance(save_path, str):
                save_path = Path(save_path)

            save_path.mkdir(parents=True, exist_ok=True)
            full_path = save_path / filename

            wb.save(str(full_path))

            ErrorHandler.log(
                f"{log_label}: {full_path}",
                level=ErrorLevel.INFO,
                context="Exportação",
            )

            return str(full_path)

        except Exception as e:
            ErrorHandler.handle_error(
                e,
                context="Exportação",
                recovery_hint="Verifique permissões de escrita no diretório de salvamento",
            )
            raise SavePathError(str(e)) from e

    def export_malote(self, malote_id: int) -> Optional[str]:
        openpyxl = _ensure_openpyxl()
        if openpyxl is None:
            return None

        registros = self._db.get_registros_with_items_by_malote(malote_id)
        if not registros:
            return None

        malote = self._db.get_malote_by_id(malote_id)
        if not malote:
            return None

        date_str = malote.date or "unknown"
        try:
            dt = datetime.fromisoformat(date_str)
            date_display = dt.strftime("%d/%m")
        except ValueError:
            date_display = date_str

        wb = openpyxl.Workbook()
        active_sheet = wb.active
        if active_sheet is not None:
            wb.remove(active_sheet)

        styles = _make_excel_styles()

        for tipo, tab_name in TIPO_LABELS.items():
            ws = wb.create_sheet(title=tab_name)

            ws["A1"] = f"USAFA OCIAN - {date_display}"
            ws.merge_cells("A1:C1")
            ws["A2"] = TIPO_TITLES[tipo]
            ws.merge_cells("A2:C2")

            tipo_registros = [r for r in registros if r.tipo == tipo]
            tipo_registros.sort(key=lambda r: r.paciente_name or "")

            for reg in tipo_registros:
                for proc in reg.processes:
                    formatted_items = [
                        _format_item(name)
                        for name in proc.items
                    ]
                    items_str = " / ".join(formatted_items)
                    return_str = ""
                    if proc.expected_return_date:
                        try:
                            dt = datetime.fromisoformat(proc.expected_return_date)
                            return_str = dt.strftime("%d/%m/%Y")
                        except (ValueError, TypeError):
                            return_str = proc.expected_return_date
                    ws.append(
                        [
                            reg.paciente_name or "",
                            items_str,
                            return_str,
                        ]
                    )

            ws.column_dimensions["A"].width = 45
            ws.column_dimensions["B"].width = 60
            ws.column_dimensions["C"].width = 15

            _style_title_row(ws, 1, styles)
            _style_title_row(ws, 2, styles, "title2_font", 26)
            _style_data_rows(ws, 3, styles)

            _apply_page_setup(ws)

        safe_date = date_display.replace("/", "-")
        return self._save_workbook(wb, "Malote", safe_date)

    def export_stats(
        self,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> Optional[str]:
        openpyxl = _ensure_openpyxl()
        if openpyxl is None:
            return None

        tipo_rows = self._db.get_stats_by_tipo(date_from=date_from, date_to=date_to)
        med_rows = self._db.get_stats_top_medications(
            date_from=date_from, date_to=date_to
        )
        if not tipo_rows and not med_rows:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Estatísticas"

        styles = _make_excel_styles()

        effective_from = date_from
        effective_to = date_to
        if not effective_from or not effective_to:
            dmin, dmax = self._db.get_malote_date_range()
            effective_from = effective_from or dmin
            effective_to = effective_to or dmax

        parts = []
        for d in (effective_from, effective_to):
            if d:
                try:
                    parts.append(datetime.fromisoformat(d).strftime("%d/%m/%Y"))
                except ValueError:
                    parts.append(d)
        date_range = " - ".join(parts) if parts else ""

        ws["A1"] = "USAFA OCIAN - Estatísticas"
        ws.merge_cells("A1:C1")

        ws["A2"] = date_range
        ws.merge_cells("A2:C2")

        ws.append(["Tipo", "Registros", "Pacientes"])
        for row in tipo_rows:
            label = TIPO_LABELS.get(row["tipo"], row["tipo"])
            ws.append([label, row["registros"], row.get("pacientes", 0)])

        totals = self._db.get_stats_totals(date_from=date_from, date_to=date_to)
        ws.append(["Total", totals["registros"], totals["pacientes"]])

        ws.append([])
        ws.append(["Medicamento", "Registros", "%"])
        total = sum(m["registros"] for m in med_rows) or 1
        for med in med_rows:
            pct = f"{med['registros'] / total * 100:.1f}%"
            ws.append(
                [
                    _format_item(med["medicamento"]),
                    med["registros"],
                    pct,
                ]
            )

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        _style_title_row(ws, 1, styles)
        _style_title_row(ws, 2, styles, "title2_font", 26)
        _style_data_rows(ws, 3, styles)

        _apply_page_setup(ws)

        date_label = ""
        if date_from or date_to:
            parts = []
            if date_from:
                try:
                    parts.append(
                        datetime.fromisoformat(date_from).strftime("%d-%m-%Y")
                    )
                except ValueError:
                    parts.append(date_from)
            parts.append("a")
            if date_to:
                try:
                    parts.append(
                        datetime.fromisoformat(date_to).strftime("%d-%m-%Y")
                    )
                except ValueError:
                    parts.append(date_to)
            date_label = "_".join(parts)
        return self._save_workbook(wb, "Estatisticas", date_label, "Estatísticas exportadas")
