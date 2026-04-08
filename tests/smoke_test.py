#!/usr/bin/env python3
"""Smoke test for RAC - tests database, state, config, export without GUI."""

import sys
import os
import tempfile
import shutil
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.database.rac_database import RACDatabase
from src.state.rac_state_manager import RACStateManager
from src.state.state_events import StateEventType, StateEvent, StateObserver
from src.utils.config import RACConfig, ConfigManager
from src.utils.text_utils import normalize_text

passed = 0
failed = 0

def test(name):
    def decorator(fn):
        def wrapper():
            global passed, failed
            try:
                fn()
                print(f"  PASS  {name}")
                passed += 1
            except Exception as e:
                print(f"  FAIL  {name}: {e}")
                failed += 1
        return wrapper
    return decorator


# ========== Text Utils ==========

@test("normalize_text removes accents")
def _():
    assert normalize_text("José Silva") == "jose silva"
    assert normalize_text("Álvaro") == "alvaro"
    assert normalize_text("") == ""
    assert normalize_text("abc") == "abc"

@test("normalize_text handles cedilha")
def _():
    assert normalize_text("França") == "franca"


# ========== Database ==========

tmp_dir = tempfile.mkdtemp(prefix="rac_test_")
db_path = os.path.join(tmp_dir, "test.db")

@test("database initializes and creates schema")
def _():
    db = RACDatabase(db_path=db_path)
    assert db.conn is not None
    db.close()

@test("catalog is seeded on init")
def _():
    db = RACDatabase(db_path=db_path)
    items = db.get_all_items()
    assert len(items) > 100, f"Expected 100+ items, got {len(items)}"
    db.close()

@test("malote CRUD")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-04-08")
    assert m["id"] is not None
    assert m["date"] == "2026-04-08"

    fetched = db.get_malote_by_id(m["id"])
    assert fetched["date"] == "2026-04-08"

    recent = db.get_recent_malotes(5)
    assert len(recent) >= 1

    all_m = db.get_all_malotes()
    assert len(all_m) >= 1
    db.close()

@test("paciente CRUD and search")
def _():
    db = RACDatabase(db_path=db_path)
    p = db.create_paciente("José da Silva")
    assert p["id"] is not None
    assert p["name"] == "José da Silva"

    fetched = db.get_paciente_by_id(p["id"])
    assert fetched["name"] == "José da Silva"

    results = db.search_pacientes("jose")
    assert len(results) >= 1

    results2 = db.search_pacientes("silva")
    assert len(results2) >= 1

    db.update_paciente(p["id"], "José Santos")
    updated = db.get_paciente_by_id(p["id"])
    assert updated["name"] == "José Santos"
    db.close()

@test("registro CRUD with items")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-05-01")
    p = db.create_paciente("Maria Oliveira")

    r = db.create_registro("entrada", p["id"], m["id"])
    assert r["id"] is not None
    assert r["tipo"] == "entrada"

    items = db.get_all_items()
    item_ids = [items[0]["id"], items[1]["id"]]
    db.set_registro_items(r["id"], item_ids)

    fetched_items = db.get_items_for_registro(r["id"])
    assert len(fetched_items) == 2

    fetched_reg = db.get_registro_by_id(r["id"])
    assert fetched_reg["paciente_name"] == "Maria Oliveira"

    regs = db.get_registros_by_malote(m["id"])
    assert len(regs) >= 1

    regs_tipo = db.get_registros_by_malote_and_tipo(m["id"], "entrada")
    assert len(regs_tipo) >= 1

    db.update_registro(r["id"], tipo="renovacao")
    updated = db.get_registro_by_id(r["id"])
    assert updated["tipo"] == "renovacao"

    assert db.delete_registro(r["id"]) == True
    assert db.get_registro_by_id(r["id"]) is None
    db.close()

@test("registro search by patient name")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-06-01")
    p1 = db.create_paciente("Ana Costa")
    p2 = db.create_paciente("Pedro Lima")

    db.create_registro("entrada", p1["id"], m["id"])
    db.create_registro("urgente", p2["id"], m["id"])

    results = db.search_registros_by_patient(m["id"], "ana")
    assert len(results) == 1
    assert results[0]["paciente_name"] == "Ana Costa"

    results2 = db.search_registros_by_patient(m["id"], "pedro")
    assert len(results2) == 1

    results3 = db.search_registros_by_patient(m["id"], "silva")
    assert len(results3) == 0
    db.close()

@test("export helper: registros with items")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-07-01")
    p1 = db.create_paciente("Beatriz Souza")
    p2 = db.create_paciente("Carlos Mendes")

    r1 = db.create_registro("entrada", p1["id"], m["id"])
    r2 = db.create_registro("entrada", p2["id"], m["id"])

    items = db.get_all_items()
    db.set_registro_items(r1["id"], [items[0]["id"], items[1]["id"]])
    db.set_registro_items(r2["id"], [items[2]["id"]])

    data = db.get_registros_with_items_by_malote(m["id"])
    assert len(data) == 2
    assert len(data[0]["items"]) == 2
    assert len(data[1]["items"]) == 1
    db.close()

@test("items catalog search")
def _():
    db = RACDatabase(db_path=db_path)
    results = db.search_items("aciclovir")
    assert len(results) >= 1
    assert "Aciclovir" in results[0]["name"]

    results2 = db.search_items("zzzznonexistent")
    assert len(results2) == 0
    db.close()

@test("malote delete: prevents delete if registros exist")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-08-01")
    p = db.create_paciente("Delete Test Patient")
    db.create_registro("entrada", p["id"], m["id"])

    assert db.delete_malote(m["id"]) == False

    db.delete_registro(db.get_registros_by_malote(m["id"])[0]["id"])
    assert db.delete_malote(m["id"]) == True
    db.close()


# ========== State Manager ==========

class TestObserver:
    def __init__(self):
        self.events = []

    def on_state_changed(self, event):
        self.events.append(event)

@test("state manager: malote events")
def _():
    sm = RACStateManager()
    obs = TestObserver()
    sm.register_observer(obs)

    assert sm.get_active_malote() is None
    assert sm.has_active_malote() == False

    sm.set_active_malote({"id": 1, "date": "2026-04-08"})
    assert sm.has_active_malote() == True
    assert sm.get_active_malote()["id"] == 1
    assert len(obs.events) == 1
    assert obs.events[0].event_type == StateEventType.MALOTE_CHANGED

@test("state manager: tipo, search, config")
def _():
    sm = RACStateManager()
    obs = TestObserver()
    sm.register_observer(obs)

    sm.set_current_tipo("entrada")
    assert sm.get_current_tipo() == "entrada"

    sm.set_search_results("jose", [{"id": 1, "name": "Jose"}])
    assert len(sm.get_search_results()) == 1

    sm.set_auto_return(False)
    assert sm.get_auto_return() == False

    sm.notify_registro_saved({"id": 1, "tipo": "entrada"})
    assert any(e.event_type == StateEventType.REGISTRO_SAVED for e in obs.events)

    sm.unregister_observer(obs)
    sm.set_current_tipo("renovacao")
    assert len([e for e in obs.events if e.event_type == StateEventType.TIPO_SELECTED]) == 1

@test("state manager: editing registro")
def _():
    sm = RACStateManager()
    assert sm.is_editing() == False

    sm.set_editing_registro({"id": 5, "tipo": "urgente"})
    assert sm.is_editing() == True
    assert sm.get_editing_registro()["id"] == 5

    sm.set_editing_registro(None)
    assert sm.is_editing() == False

@test("state manager: deep copy on get")
def _():
    sm = RACStateManager()
    sm.set_active_malote({"id": 1, "date": "2026-01-01"})

    m = sm.get_active_malote()
    m["id"] = 999

    assert sm.get_active_malote()["id"] == 1


# ========== Config ==========

@test("RACConfig defaults")
def _():
    c = RACConfig()
    assert c.auto_return == True
    assert c.theme == "dark"
    assert c.last_malote_id is None

@test("RACConfig validation")
def _():
    c = RACConfig(auto_return=False, theme="light")
    assert c.auto_return == False
    assert c.theme == "light"

    try:
        RACConfig(theme="invalid")
        assert False, "Should have raised ValueError"
    except ValueError:
        pass


# ========== Export ==========

@test("excel export produces file")
def _():
    db = RACDatabase(db_path=db_path)
    m = db.create_malote("2026-09-01")
    p1 = db.create_paciente("Export Patient 1")
    p2 = db.create_paciente("Export Patient 2")

    r1 = db.create_registro("entrada", p1["id"], m["id"])
    r2 = db.create_registro("renovacao", p2["id"], m["id"])

    items = db.get_all_items()
    db.set_registro_items(r1["id"], [items[0]["id"]])
    db.set_registro_items(r2["id"], [items[1]["id"], items[2]["id"]])

    export_dir = os.path.join(tmp_dir, "exports")
    os.makedirs(export_dir, exist_ok=True)

    from src.export.excel_exporter import ExcelExporter
    exporter = ExcelExporter(db)

    result = exporter.export_malote(m["id"])
    assert result is not None
    assert os.path.exists(result)
    assert result.endswith(".xlsx")

    import openpyxl
    wb = openpyxl.load_workbook(result)
    assert "Entradas" in wb.sheetnames
    assert "Renovações" in wb.sheetnames
    assert "Retiradas" in wb.sheetnames
    assert "Urgentes" in wb.sheetnames

    ws_entrada = wb["Entradas"]
    assert ws_entrada.cell(row=1, column=1).value == "Nome"
    assert ws_entrada.cell(row=1, column=2).value == "Medicamentos"
    assert ws_entrada.cell(row=2, column=1).value == "Export Patient 1"

    wb.close()
    db.close()


# ========== Run ==========

if __name__ == "__main__":
    print("\n=== RAC Smoke Tests ===\n")

    # Collect and run all tests
    tests = []
    frame = sys._getframe()
    for name, obj in list(globals().items()):
        if callable(obj) and hasattr(obj, '__wrapped__'):
            tests.append(obj)

    for t in tests:
        t()

    print(f"\n{'='*40}")
    print(f"Results: {passed} passed, {failed} failed, {passed + failed} total")
    print(f"{'='*40}\n")

    # Cleanup
    shutil.rmtree(tmp_dir, ignore_errors=True)

    sys.exit(1 if failed > 0 else 0)
